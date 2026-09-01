import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

COUNT_PATTERN = re.compile(r"^(.*?)(?:\s+x?(\d+))?$", re.IGNORECASE)
STANDARD_PATTERN = re.compile(r"^(\d+)([RW]*)$")
STANDARD_SPECIAL_PATTERN = re.compile(r"^(\d+)([RW]*)\s*\((RH|LBA)\s+(\d+)\)$")
AQUARIUM_PATTERN = re.compile(r"^\((\d+)\)\s*Aq\s+(\d+)\s*([RW]*)$")
AQUARIUM_ALTERNATIVE_PATTERN = re.compile(
    r"^\((\d+)\)\s*Aq\s+(\d+)\s*/\s*(LRH)\s+(\d+)$"
)
STANDARD_AQUARIUM_PATTERN = re.compile(r"^(\d+)([RW]*)\s*/\s*Aq\s+(\d+)$")
SPECIAL_ONLY_PATTERN = re.compile(r"^(PZ)\s+(\d+)$")


def _clean_label(value: str) -> str:
    return re.sub(r"\s+", " ", value.strip()).lower()


def _counted_values(value: Any, separator: str) -> list[str]:
    if value is None or not str(value).strip():
        return []
    values: list[str] = []
    for part in re.split(separator, str(value).strip()):
        match = COUNT_PATTERN.fullmatch(part.strip())
        if not match:
            raise ValueError(f"Could not parse counted value: {part!r}")
        label = _clean_label(match.group(1))
        count = int(match.group(2) or 1)
        values.extend([label] * count)
    return values


def parse_types(value: Any) -> list[str]:
    """Return all printed animal-category icons, retaining duplicate icons."""
    return _counted_values(value, r"\s*/\s*")


def parse_requirements(value: Any) -> dict[str, int]:
    """Return requirements as normalized labels and required counts."""
    requirements: dict[str, int] = {}
    for requirement in _counted_values(value, r"[\r\n]+"):
        requirements[requirement] = requirements.get(requirement, 0) + 1
    return requirements


def parse_bonuses(value: Any) -> dict[str, int]:
    parts = [int(part.strip()) for part in str(value).split("/")]
    if len(parts) != 3:
        raise ValueError(f"Expected bonuses in A/C/R format, got {value!r}")
    return {"appeal": parts[0], "conservation": parts[1], "reputation": parts[2]}


def _special(code: str, spaces: int) -> dict[str, Any]:
    return {
        "type": "aquarium" if code.lower() == "aq" else code.lower(),
        "spaces": spaces,
    }


def parse_enclosure(value: Any) -> dict[str, Any]:
    """Parse standard-enclosure conditions and all printed special alternatives."""
    text = re.sub(r"\s+", " ", str(value).strip())
    size: int | None = None
    conditions = ""
    special_enclosures: list[dict[str, Any]] = []

    if match := STANDARD_PATTERN.fullmatch(text):
        size, conditions = int(match.group(1)), match.group(2)
    elif match := STANDARD_SPECIAL_PATTERN.fullmatch(text):
        size, conditions = int(match.group(1)), match.group(2)
        special_enclosures.append(_special(match.group(3), int(match.group(4))))
    elif match := AQUARIUM_PATTERN.fullmatch(text):
        size, conditions = int(match.group(1)), match.group(3)
        special_enclosures.append(_special("Aq", int(match.group(2))))
    elif match := AQUARIUM_ALTERNATIVE_PATTERN.fullmatch(text):
        size = int(match.group(1))
        special_enclosures.extend(
            [
                _special("Aq", int(match.group(2))),
                _special(match.group(3), int(match.group(4))),
            ]
        )
    elif match := STANDARD_AQUARIUM_PATTERN.fullmatch(text):
        size, conditions = int(match.group(1)), match.group(2)
        special_enclosures.append(_special("Aq", int(match.group(3))))
    elif match := SPECIAL_ONLY_PATTERN.fullmatch(text):
        special_enclosures.append(_special(match.group(1), int(match.group(2))))
    else:
        raise ValueError(f"Unknown enclosure notation: {value!r}")

    return {
        "size": size,
        "water": conditions.count("W"),
        "rock": conditions.count("R"),
        "special_enclosure": special_enclosures,
    }


def parse_animal_workbook(path: Path) -> list[dict[str, Any]]:
    """Convert the workbook's Animals sheet to normalized dictionaries."""
    workbook = load_workbook(path, data_only=True, read_only=True)
    worksheet = workbook["Animals"]
    rows = worksheet.iter_rows(values_only=True)
    headers = next(rows)
    cards = []
    for values in rows:
        source = dict(zip(headers, values, strict=False))
        if source["Card #"] is None:
            continue
        cards.append(
            {
                "card_number": int(source["Card #"]),
                "name": str(source["Animal Card Name"]).strip().title(),
                "latin_name": source["Animal Latin name"],
                "cost": int(source["Cost"]),
                "types": parse_types(source["Type"]),
                "continent": _clean_label(str(source["Continent"])),
                "requirements": parse_requirements(source["Reqs"]),
                "bonuses": parse_bonuses(source["Bonuses (A/C/R)"]),
                "enclosure": parse_enclosure(source["Enclosure size (Rock/Water)"]),
                "ability": source["Ability"],
                "reef_ability": source["Reef Ability"],
                "wave_icon": bool(source["Wave Icon"]),
                "expansion": source["Expansion (MW)"],
            }
        )
    return cards
